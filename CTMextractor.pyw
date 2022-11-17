import keyboard
import win32gui
import win32ui
import os
import time
import pyautogui
from pyWinActivate import win_activate, win_wait_active, get_app_list
import pandas as pd
import clipboard
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment



def load_aliases():
    global alias
    alias = {}

    file = open('distalias.txt', 'r')
    alias_file = file.readlines()
    file.close()


    for line in alias_file:
        l = line.split(sep="=")
        alias[l[0]] = l[1].strip("\n")

def load_special_cases():
    global special
    special = []

    file = open('distspecial.txt', 'r')
    special_file = file.readlines()
    file.close()

    for line in special_file:
        special.append(line)

def load_ignore_list():
    global ignore
    ignore = []
    file = open('distignore.txt', 'r')
    ignore_file = file.readlines()
    file.close()

    for line in ignore_file:
        ignore.append(line)

def load_countries_filter():
    global countries_filter
    countries_filter = {}

    file = open('countries.txt', 'r')
    filter_file = file.readlines()
    file.close()


    for line in filter_file:
        l = line.split(sep="=")
        countries_filter[l[0]] = l[1].strip("\n")


def ctm(x1, y1, x2, y2, num_of_loops):

    def terminate():
        if keyboard.is_pressed('end'):
            pyautogui.alert(text='Process was force stopped.', title='Alert! CCT data downloader', button='OK')
            quit()

    """ ALL POSSIBLE WINDOWS'S TITLES
    Distribution Panel CCT @GfK - [dip.gfk.com][01.02.07] - \\\\Remote
    Distribution Panel CCT *ERROR - \\\\Remote
    TDistribution Panel CCT *WARNING - \\\\Remote
    Loading... - \\\\Remote
    """


    title = "Distribution Panel CCT @GfK - [dip.gfk.com]"
    no_records_found = "TDistribution Panel CCT *WARNING - \\\\Remote"
    change_country_error = "Distribution Panel CCT *ERROR - \\\\Remote"

    total_countries = num_of_loops #total: 73 | 33

    data_list = []


    load_aliases()

    load_special_cases()

    load_ignore_list()

    load_countries_filter()


    def sleep(sec=0.5):
        time.sleep(sec)

    def press(key=str()):
        keyboard.press_and_release(key)

    def get_country_data():
        change_country()
        
        try:
            error_check(title=no_records_found)
        except:
            try:
                error_check(title=change_country_error)
            except:
                pass
            copy_data()

    def change_country():
        terminate()
        pyautogui.click(x=x1 ,y=y1)
        sleep()
        press('down')
        sleep()
        press('enter')
        sleep()
        press('tab')
        sleep()
        press('tab')
        sleep()
        press('tab')
        sleep()
        press('enter')
        sleep()
        win_wait_active(win_to_wait=title, exception=no_records_found, message=False)
        sleep()

        terminate()
        # check and clear error if it exists
        try:
            error_check(title=change_country_error)
        except:
            print("No error window found. No actions required.")

        #focus back on main window
        win_activate(window_title=title, partial_match=True)
        win_wait_active(win_to_wait=title, message=False)

    def copy_data():
        terminate()
        sleep()
        pyautogui.click(x=x2 ,y=y2)
        sleep()
        keyboard.send('ctrl+a')
        sleep()
        # copy and add to dataframe
        keyboard.send('ctrl+c') #shows windows sound menu on copy
        sleep(3)
        data = pd.read_clipboard(sep='\t')
        sleep(1)

        #evaluate if the current copied data's country is true or false in the countries filter list
        # if its true, add it to the final data_list, otherwise skip it 



        try:
            if (countries_filter[data.iloc[1,1]]) == "True":
                data_list.append(data)
        except:
            if (countries_filter[data.iloc[0,1]]) == "True":
                data_list.append(data) 

        sleep()
        

    def error_check(title=str()):
            win_activate(window_title=title)
            win_wait_active(win_to_wait=title, message=False)
            sleep()
            press('esc')



    '''
    MAINLOOP
    '''
    win_activate(window_title=title, partial_match=True)
    win_wait_active(win_to_wait=title, message=False)
    for i in range(total_countries):
        terminate()
        get_country_data()


    '''
    Data file processing
    '''

    df = pd.concat(data_list)

    #remove rows that are duplicates of the header
    df = df[df.Period != "Period"]
    #create new columns
    df['New'] = 1
    df['Done'] = 1
    df.rename(columns={'Client Type': 'Client_Type'}, inplace=True)


    #count all unique distributors
    new = df.groupby(['Period', 'ISO', 'Distributor Name']).count()['New']

    #evaluate dummys as 0
    df.loc[(df.Client_Type == "DUMMY"), "Done"] = 0

    #sum all done records
    new2 = df.groupby(['Period', 'ISO', 'Distributor Name']).sum()['Done']

    #convert series new2 to dataframe
    df2 = new2.to_frame().reset_index()
    df2 = df2.rename(columns= {0: 'Done'})
    #extract only the column i want
    df2 = df2['Done']


    # save DataFrame to excel file
    writer = pd.ExcelWriter('PeriodData.xlsx')
    new.to_excel(writer, "data", merge_cells=False)
    df2.to_excel(writer, "data" ,startcol=4, index=False, merge_cells=False)
    writer.save()


    # open with excel editor
    wb = openpyxl.load_workbook(filename='PeriodData.xlsx')
    ws = wb.active
    # delete 0 values from "Done" column

    for cell in ws['E:E']:
        if cell.value == 0:
            cell.value = None

    # add new column "Pending" + formula
    ws['F1'].value = "Pending"
    ws['F2'].value = "=D2-E2"
    # add new column "FTE"
    ws['G1'].value = "FTE"
    ws['G2'].value = 1
    # add new column "CT per FTE" + formula
    ws['H1'].value = "CT per FTE"
    ws['H2'].value = "=E2/G2"

    #convert period formatting
    per = ws['A2']
    per.value = "20%s-%s" % (str(per.value)[0:2], str(per.value)[2:4])
    perstr = str(ws['A2'].value)

    #delete ignore cases
    for cell in ws['C']:
        if cell.value in ignore:
            ws.delete_rows(idx=cell.row , amount=1)

    #unmerge period cell and fill to end of column
    a_last = len(ws['A'])
    # ws.unmerge_cells('A2:A%i' % (a_last)) OBSOLETE (ROW 58)



    #A
    for i in range(3, a_last+1):
        ws.cell(row=i, column=1).value = ws['A2'].value
    #F
    for i in range(3, a_last+1):
        ws.cell(row=i, column=6).value = "=D%s-E%s" % (str(i), str(i))
    #G
    for i in range(3, a_last+1):
        ws.cell(row=i, column=7).value = 1
    #H
    for i in range(3, a_last+1):
        ws.cell(row=i, column=8).value = "=E%s/G%s" % (str(i), str(i))





    # change table style
    wsf = ws['E5']
    font = Font(
        name= 'Calibri',
        size=11,
        color='00969696'
    )
    side = Side(border_style=None)
    no_border = Border(
        left=side, 
        right=side, 
        top=side, 
        bottom=side,
    )
    alignment = Alignment(
        horizontal='general',
        vertical='bottom'
    )
    ali_c_header = Alignment(horizontal='center')
    ali_a = Alignment(horizontal='center')

    for cell in ws['A']:
        cell.font = font
        cell.border = no_border
        cell.alignment = ali_a
    for cell in ws['B']:
        cell.font = font
        cell.border = no_border
    for cell in ws['C']:
        cell.font = font
        cell.border = no_border
        cell.alignment = alignment
    cell = ws['C1']
    cell.alignment = ali_c_header
    for cell in ws['D']:
        cell.font = font
        cell.border = no_border
    for cell in ws['E']:
        cell.font = font
        cell.border = no_border
    for cell in ws['F']:
        cell.font = font
    for cell in ws['G']:
        cell.font = font
    for cell in ws['H']:
        cell.font = font



    #recolor special cases
    special_font = Font(
        color='00800000'
    )
    special_fill = PatternFill(
        start_color='00FF99CC',
        end_color='00FF99CC',
        fill_type='solid'
    )


    for cell in ws['C']:
        if cell.value in special:
            cell.font = special_font
            cell.fill = special_fill

    #rename all distributor names with aliases
    for cell in ws['C']:
        if cell.value in alias:
            cell.value = alias[cell.value]



    ws.column_dimensions['C'].width = 26.71
    wb.save('PeriodData.xlsx')
    wb.close()


    # display message box when finished
    pyautogui.alert(text='Downloading data from CCT has finished.', title='Alert! CCT data downloader', button='OK')