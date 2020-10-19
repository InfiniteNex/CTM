import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os

#load aliases
alias = {}

file = open('distalias.txt', 'r')
alias_file = file.readlines()
file.close()


for line in alias_file:
    l = line.split(sep="=")
    alias[l[0]] = l[1]


#load special cases
special = []

file = open('distspecial.txt', 'r')
special_file = file.readlines()
file.close()

for line in special_file:
    special.append(line)


#read data file
df = pd.read_excel('CTMdataStorage.xlsx')
#remove rows that are duplicates of the header
df = df[df.Period != "Period"]
#create new columns
df['New'] = 1
df['Done'] = 1
df.rename(columns={'Client Type': 'Client_Type'}, inplace=True)


#count all unique distributors
new = df.groupby(['Period', 'ISO', 'Distributor Name']).count()['New']

#evaluate dummys as 0
df.loc[(df.Client_Type == "DUMMY"), "Done"] = 0

#sum all done records
new2 = df.groupby(['Period', 'ISO', 'Distributor Name']).sum()['Done']

#convert series new2 to dataframe
df2 = new2.to_frame().reset_index()
df2 = df2.rename(columns= {0: 'Done'})
#extract only the column i want
df2 = df2['Done']


# save DataFrame to excel file
writer = pd.ExcelWriter('result.xlsx')
new.to_excel(writer, "data", merge_cells=False)
df2.to_excel(writer, "data" ,startcol=4, index=False, merge_cells=False)
writer.save()


# open with excel editor
wb = openpyxl.load_workbook(filename='result.xlsx')
ws = wb.active
# delete 0 values from "Done" column

for cell in ws['E:E']:
    if cell.value == 0:
        cell.value = None

# add new column "Pending" + formula
ws['F1'].value = "Pending"
ws['F2'].value = "=D2-E2"
# add new column "FTE"
ws['G1'].value = "FTE"
ws['G2'].value = 1
# add new column "CT per FTE" + formula
ws['H1'].value = "CT per FTE"
ws['H2'].value = "=E2/G2"

#convert period formatting
per = ws['A2']
per.value = "20%s-%s" % (str(per.value)[0:2], str(per.value)[2:4])
perstr = str(ws['A2'].value)

#unmerge period cell and fill to end of column
a_last = len(ws['A'])
# ws.unmerge_cells('A2:A%i' % (a_last)) OBSOLETE (ROW 58)

#A
for i in range(3, a_last+1):
    ws.cell(row=i, column=1).value = ws['A2'].value
#F
for i in range(3, a_last+1):
    ws.cell(row=i, column=6).value = "=D%s-E%s" % (str(i), str(i))
#G
for i in range(3, a_last+1):
    ws.cell(row=i, column=7).value = 1
#H
for i in range(3, a_last+1):
    ws.cell(row=i, column=8).value = "=E%s/G%s" % (str(i), str(i))





# change table style
wsf = ws['E5']
font = Font(
    name= 'Calibri',
    size=11,
    color='00969696'
)
side = Side(border_style=None)
no_border = Border(
    left=side, 
    right=side, 
    top=side, 
    bottom=side,
)
alignment = Alignment(
    horizontal='general',
    vertical='bottom'
)
ali_c_header = Alignment(horizontal='center')
ali_a = Alignment(horizontal='center')

for cell in ws['A']:
    cell.font = font
    cell.border = no_border
    cell.alignment = ali_a
for cell in ws['B']:
    cell.font = font
    cell.border = no_border
for cell in ws['C']:
    cell.font = font
    cell.border = no_border
    cell.alignment = alignment
cell = ws['C1']
cell.alignment = ali_c_header
for cell in ws['D']:
    cell.font = font
    cell.border = no_border
for cell in ws['E']:
    cell.font = font
    cell.border = no_border
for cell in ws['F']:
    cell.font = font
for cell in ws['G']:
    cell.font = font
for cell in ws['H']:
    cell.font = font



#recolor special cases
special_font = Font(
    color='00800000'
)
special_fill = PatternFill(
    start_color='00FF99CC',
    end_color='00FF99CC',
    fill_type='solid'
)


for cell in ws['C']:
    if cell.value in special:
        cell.font = special_font
        cell.fill = special_fill

#rename all distributor names with aliases
for cell in ws['C']:
    if cell.value in alias:
        cell.value = alias[cell.value]



ws.column_dimensions['C'].width = 26.71
wb.save('result.xlsx')
wb.close()


#empty CTMdataStorage.xlsx for future use
wb = openpyxl.load_workbook(filename='CTMdataStorage.xlsx')
ws = wb.active
ws.delete_cols(idx=1, amount=11)
wb.save('CTMdataStorage.xlsx')
wb.close()